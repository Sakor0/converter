"""
data.py - Konwersje plików danych: CSV, JSON, XLSX, YAML. Czysty Python.
"""

import csv
import json

import openpyxl
import yaml

DATA_EXTS = {".csv", ".json", ".xlsx", ".yaml", ".yml"}


def _read_csv_rows(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def csv_to_json(input_path, output_path):
    rows = _read_csv_rows(input_path)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def json_to_csv(input_path, output_path):
    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = [data]
    fieldnames = sorted({k for row in data for k in row.keys()})
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)


def csv_to_xlsx(input_path, output_path):
    rows = _read_csv_rows(input_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    if rows:
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
    wb.save(output_path)


def xlsx_to_csv(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)


def json_to_yaml(input_path, output_path):
    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)
    with open(output_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)


def yaml_to_json(input_path, output_path):
    with open(input_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# Rejestr bezpośrednich ścieżek konwersji - używany przez convert.py
CONVERSIONS = {
    (".csv", ".json"): csv_to_json,
    (".json", ".csv"): json_to_csv,
    (".csv", ".xlsx"): csv_to_xlsx,
    (".xlsx", ".csv"): xlsx_to_csv,
    (".json", ".yaml"): json_to_yaml,
    (".json", ".yml"): json_to_yaml,
    (".yaml", ".json"): yaml_to_json,
    (".yml", ".json"): yaml_to_json,
}
